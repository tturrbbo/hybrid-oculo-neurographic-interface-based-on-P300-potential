"""Manual export of recorded P300 runs to txt/csv/xlsx."""

from __future__ import annotations

import csv
import math
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np

from p300_analysis.marker_parsing import (
    decode_stim_tile_id,
    marker_value_to_stim_key,
    parse_trial_target_tile_id,
    stim_key_to_tile_digit,
)


def _parse_tile_event(value: Any) -> Optional[Tuple[int, str]]:
    """Разбирает маркер стима плитки в пару ``(tile_digit, phase)``, где ``phase`` — ``"on"`` или ``"off"``.

    Принимает строки вида ``"5|on"``, ``"5|off"``. Для ``trial_start/trial_end``
    и отрицательных id возвращает ``None``.
    """
    if isinstance(value, (bytes, bytearray)):
        try:
            value = value.decode("utf-8", errors="ignore")
        except Exception:
            return None
    if not isinstance(value, str):
        return None
    s = value.strip()
    if "|" not in s:
        return None
    left, right = s.split("|", 1)
    try:
        raw_id = int(left.strip())
    except ValueError:
        return None
    if raw_id < 0:
        return None
    tile = decode_stim_tile_id(raw_id)
    if tile is None:
        return None
    phase = right.split("|", 1)[0].strip().lower()
    if phase not in {"on", "off"}:
        return None
    return tile, phase


def _ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _round3(value: Any) -> Any:
    if isinstance(value, bool) or value is None:
        return value
    if isinstance(value, (int, float)):
        fv = float(value)
        if not math.isfinite(fv):
            return None
        return round(fv, 3)
    return value


def _rounded_row(row: Sequence[Any]) -> List[Any]:
    return [_round3(x) for x in row]


def _rows_to_csv(path: Path, header: Sequence[str], rows: Iterable[Sequence[Any]]) -> None:
    _ensure_parent(path)
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(list(header))
        for row in rows:
            writer.writerow(_rounded_row(row))


def _format_ru_decimal(value: Any) -> Any:
    """Для русской локали Excel: число → строка с запятой вместо точки."""
    if isinstance(value, bool) or value is None:
        return value
    if isinstance(value, (int, np.integer)):
        return int(value)
    if isinstance(value, (float, np.floating)):
        fv = float(value)
        if not math.isfinite(fv):
            return ""
        return (f"{round(fv, 6):.6f}").rstrip("0").rstrip(".").replace(".", ",") or "0"
    return value


def _rows_to_csv_ru(
    path: Path,
    header: Sequence[str],
    rows: Iterable[Sequence[Any]],
) -> None:
    """CSV в «русском» формате Excel: разделитель столбцов ``;``, десятичный — ``,``.
    В первой строке пишется ``sep=;``, чтобы Excel корректно определил разделитель.
    """
    _ensure_parent(path)
    with open(path, "w", encoding="utf-8", newline="") as f:
        f.write("sep=;\n")
        writer = csv.writer(f, delimiter=";")
        writer.writerow(list(header))
        for row in rows:
            writer.writerow([_format_ru_decimal(x) for x in row])


def _rows_to_xlsx(
    path: Path,
    sheet_name: str,
    header: Sequence[str],
    rows: Iterable[Sequence[Any]],
) -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:  # pragma: no cover - runtime dependency check
        raise RuntimeError(
            "Для экспорта в XLSX установите openpyxl (pip install openpyxl)."
        ) from e
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "data"
    ws.append(list(header))
    for row in rows:
        ws.append(_rounded_row(row))
    _ensure_parent(path)
    wb.save(path)


def _rows_to_txt(path: Path, title: str, header: Sequence[str], rows: Iterable[Sequence[Any]]) -> None:
    _ensure_parent(path)
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"{title}\n")
        f.write("\t".join(header) + "\n")
        for row in rows:
            f.write("\t".join(str(x) for x in _rounded_row(row)) + "\n")


def _summary_rows(run_data: Dict[str, Any]) -> List[Tuple[str, Any]]:
    summary = run_data.get("summary") or {}
    params = summary.get("analysis_params") or {}
    return [
        ("run_seq", run_data.get("run_seq")),
        ("saved_at_ms", run_data.get("saved_at_ms")),
        ("n_markers", len(run_data.get("markers") or [])),
        ("n_eeg_samples", len(run_data.get("eeg_ts") or [])),
        ("n_winner_updates", len(run_data.get("winner_updates") or [])),
        ("n_epoch_classes", len(run_data.get("epochs_data") or {})),
        ("baseline_ms", params.get("baseline_ms")),
        ("window_x_ms", params.get("window_x_ms")),
        ("window_y_ms", params.get("window_y_ms")),
        ("epochs_after_trial_only", params.get("epochs_after_trial_only")),
        ("ui_winner_tile_id", summary.get("ui_winner_tile_id")),
        ("last_lsl_cue", summary.get("last_lsl_cue")),
        ("match_last_cue_vs_winner", summary.get("match_last_cue_vs_winner")),
    ]


def _marker_rows(run_data: Dict[str, Any]) -> List[Tuple[Any, Any]]:
    rows: List[Tuple[Any, Any]] = []
    for item in run_data.get("markers") or []:
        rows.append((item.get("ts"), item.get("value")))
    return rows


def _eeg_rows(run_data: Dict[str, Any]) -> List[List[Any]]:
    rows: List[List[Any]] = []
    eeg_ts = run_data.get("eeg_ts") or []
    eeg_samples = run_data.get("eeg_samples") or []
    for i, (ts, sample) in enumerate(zip(eeg_ts, eeg_samples)):
        row: List[Any] = [i, ts]
        row.extend(sample if isinstance(sample, list) else [sample])
        rows.append(row)
    return rows


def _winner_rows(run_data: Dict[str, Any]) -> List[Tuple[Any, Any, Any, Any]]:
    rows: List[Tuple[Any, Any, Any, Any]] = []
    for item in run_data.get("winner_updates") or []:
        rows.append(
            (
                item.get("event_seq"),
                item.get("winner_digit"),
                item.get("winner_key"),
                item.get("match_lsl_cue"),
            )
        )
    return rows


def _epoch_rows(run_data: Dict[str, Any]) -> List[Tuple[Any, Any, Any, Any]]:
    rows: List[Tuple[Any, Any, Any, Any]] = []
    time_ms = run_data.get("epoch_time_ms") or []
    epochs_data = run_data.get("epochs_data") or {}
    for stim_key, epochs in epochs_data.items():
        for epoch_idx, epoch_values in enumerate(epochs):
            for sample_idx, value in enumerate(epoch_values):
                t_ms = time_ms[sample_idx] if sample_idx < len(time_ms) else None
                rows.append((stim_key, epoch_idx, t_ms, value))
    return rows


def _t_rel_marker_ms(run_data: Dict[str, Any], sample_idx: int) -> Optional[float]:
    """Мс от вспышки (0 = маркер); не зависит от грубых штампов потока ЭЭГ."""
    tplate = run_data.get("epoch_time_ms") or []
    if sample_idx < len(tplate):
        return float(tplate[sample_idx])
    ap = (run_data.get("summary") or {}).get("analysis_params") or {}
    srate = ap.get("sampling_rate_hz")
    if srate is not None and float(srate) > 0:
        return float(sample_idx) * (1000.0 / float(srate))
    return None


def _t_sync_est_marker_axis(marker_ts: Any, rel_ms: Optional[float]) -> Optional[float]:
    """Оценка времени отсчёта в той же оси, что ``marker_ts`` (LSL): marker + rel_ms."""
    if rel_ms is None:
        return None
    try:
        return float(marker_ts) + float(rel_ms) / 1000.0
    except (TypeError, ValueError):
        return None


# Служебных колонок перед ch_* в экспорте одной плитки (расширенный ряд для отладки).
_EPOCH_EXPORT_META_COLS = 7
# Служебных колонок в экспорте «все плитки» (как в UI NeuroSpectrum: segment_id, …, ts).
_ALL_STIMS_EXPORT_META_COLS = 5


def _epoch_raw_rows_all_stims_blink_order(run_data: Dict[str, Any]) -> List[List[Any]]:
    """Все эпохи подряд в порядке вспышек (``marker_ts``), ``segment_id`` = 0,1,2… по хронологии."""
    segments = list(run_data.get("epoch_segments") or [])
    indexed = list(enumerate(segments))
    indexed.sort(key=lambda p: (float(p[1].get("marker_ts") or 0.0), p[0]))
    rows: List[List[Any]] = []
    for blink_ord, (_orig_i, seg) in enumerate(indexed):
        stim_key = seg.get("stim_key")
        marker_ts = seg.get("marker_ts")
        eeg_ts = seg.get("eeg_ts") or []
        eeg_samples = seg.get("eeg_samples") or []
        for sample_idx, (ts, sample) in enumerate(zip(eeg_ts, eeg_samples)):
            row: List[Any] = [blink_ord, stim_key, marker_ts, sample_idx, ts]
            row.extend(sample if isinstance(sample, list) else [sample])
            rows.append(row)
    return rows


def _stim_index_from_key(stim_key: Any) -> Optional[int]:
    if not isinstance(stim_key, str):
        return None
    if "_" not in stim_key:
        return None
    tail = stim_key.rsplit("_", 1)[-1]
    try:
        return int(tail)
    except Exception:
        return None


def stim_indices_in_run(run_data: Dict[str, Any]) -> List[int]:
    """Sorted unique stim indices present in ``epoch_segments`` (e.g. 0..8)."""
    out: set[int] = set()
    for seg in run_data.get("epoch_segments") or []:
        idx = _stim_index_from_key(seg.get("stim_key"))
        if idx is not None:
            out.add(idx)
    return sorted(out)


def _epoch_raw_rows_for_stim(run_data: Dict[str, Any], stim_index: int) -> List[List[Any]]:
    rows: List[List[Any]] = []
    segments = run_data.get("epoch_segments") or []
    for seg_idx, seg in enumerate(segments):
        stim_key = seg.get("stim_key")
        if _stim_index_from_key(stim_key) != stim_index:
            continue
        marker_ts = seg.get("marker_ts")
        eeg_ts = seg.get("eeg_ts") or []
        eeg_samples = seg.get("eeg_samples") or []
        for sample_idx, (ts, sample) in enumerate(zip(eeg_ts, eeg_samples)):
            rel_ms = _t_rel_marker_ms(run_data, sample_idx)
            t_sync = _t_sync_est_marker_axis(marker_ts, rel_ms)
            row: List[Any] = [seg_idx, stim_key, marker_ts, sample_idx, rel_ms, t_sync, ts]
            row.extend(sample if isinstance(sample, list) else [sample])
            rows.append(row)
    return rows


def _filter_sample_channels(sample: Any, channels: List[int]) -> List[Any]:
    vals = sample if isinstance(sample, list) else [sample]
    out: List[Any] = []
    for c in channels:
        if 0 <= c < len(vals):
            out.append(vals[c])
    return out


def _filtered_run_data(run_data: Dict[str, Any], selected_channels: List[int] | None) -> Dict[str, Any]:
    if selected_channels is None:
        return run_data
    if not selected_channels:
        raise RuntimeError("Список каналов для экспорта пустой.")
    filtered = dict(run_data)
    filtered["eeg_samples"] = [
        _filter_sample_channels(sample, selected_channels) for sample in (run_data.get("eeg_samples") or [])
    ]
    segs: List[Dict[str, Any]] = []
    for seg in run_data.get("epoch_segments") or []:
        seg_copy = dict(seg)
        seg_copy["eeg_samples"] = [
            _filter_sample_channels(sample, selected_channels) for sample in (seg.get("eeg_samples") or [])
        ]
        segs.append(seg_copy)
    filtered["epoch_segments"] = segs
    filtered["selected_channels"] = list(selected_channels)
    return filtered


def export_run_data(
    *,
    run_data: Dict[str, Any],
    output_path: Path,
    file_format: str,
    stim_index: int,
    selected_channels: List[int] | None = None,
) -> List[Path]:
    """Export EEG rows only for selected stimulus index."""
    run_data = _filtered_run_data(run_data, selected_channels)
    file_format = file_format.lower().strip()
    created: List[Path] = []
    stem = output_path.with_suffix("")

    rows = _epoch_raw_rows_for_stim(run_data, stim_index=stim_index)
    if not rows:
        raise RuntimeError(f"Нет эпох для stim_index={stim_index}.")
    max_ch = max((len(x) for x in (run_data.get("eeg_samples") or [])), default=0)
    if max_ch == 0:
        max_ch = max((len(r) - _EPOCH_EXPORT_META_COLS for r in rows), default=0)
    header = [
        "segment_idx",
        "stim_key",
        "marker_ts",
        "sample_idx",
        "t_rel_marker_ms",
        "t_sync_est",
        "eeg_ts_stream",
    ] + [f"ch_{i+1}" for i in range(max_ch)]

    if file_format in {"csv", "txt"}:
        ext = ".csv" if file_format == "csv" else ".txt"
        p = Path(f"{stem}_stim_{stim_index}{ext}")
        if file_format == "csv":
            _rows_to_csv(p, header, rows)
        else:
            _rows_to_txt(p, f"EEG at flashes for stim {stim_index}", header, rows)
        created.append(p)
        return created

    if file_format == "xlsx":
        try:
            from openpyxl import Workbook
        except Exception as e:  # pragma: no cover - runtime dependency check
            raise RuntimeError(
                "Для экспорта в XLSX установите openpyxl (pip install openpyxl)."
            ) from e

        wb = Workbook()
        wb.remove(wb.active)

        def _add_sheet(name: str, header: Sequence[str], rows: Iterable[Sequence[Any]]) -> None:
            ws = wb.create_sheet(title=name[:31])
            ws.append(list(header))
            for row in rows:
                ws.append(_rounded_row(row))

        _add_sheet(f"stim_{stim_index}", header, rows)

        final_path = output_path if output_path.suffix.lower() == ".xlsx" else output_path.with_suffix(".xlsx")
        _ensure_parent(final_path)
        wb.save(final_path)
        created.append(final_path)
        return created

    raise ValueError(f"Unsupported file format: {file_format}")


def export_run_data_all_stims(
    *,
    run_data: Dict[str, Any],
    output_path: Path,
    file_format: str,
    selected_channels: List[int] | None = None,
) -> List[Path]:
    """Export raw EEG for every flash into one file.

    Эпохи идут **в порядке вспышек** (сортировка по ``marker_ts``). Колонки:
    ``segment_id`` (0… по времени), ``stim_key``, ``marker_ts``, ``sample_idx``, ``ts``, каналы.
    """
    run_data = _filtered_run_data(run_data, selected_channels)
    file_format = file_format.lower().strip()
    created: List[Path] = []

    rows = _epoch_raw_rows_all_stims_blink_order(run_data)
    if not rows:
        raise RuntimeError("Нет сегментов epoch_segments — нечего экспортировать.")

    max_ch = max((len(x) for x in (run_data.get("eeg_samples") or [])), default=0)
    if max_ch == 0:
        max_ch = max((len(r) - _ALL_STIMS_EXPORT_META_COLS for r in rows), default=0)
    header = ["segment_id", "stim_key", "marker_ts", "sample_idx", "ts"] + [
        f"ch_{i+1}" for i in range(max_ch)
    ]

    if file_format in {"csv", "txt"}:
        ext = ".csv" if file_format == "csv" else ".txt"
        p = output_path if output_path.suffix.lower() == ext else output_path.with_suffix(ext)
        if file_format == "csv":
            _rows_to_csv(p, header, rows)
        else:
            _rows_to_txt(p, "EEG at flashes (all stims)", header, rows)
        created.append(p)
        return created

    if file_format == "xlsx":
        try:
            from openpyxl import Workbook
        except Exception as e:  # pragma: no cover - runtime dependency check
            raise RuntimeError(
                "Для экспорта в XLSX установите openpyxl (pip install openpyxl)."
            ) from e

        wb = Workbook()
        wb.remove(wb.active)

        def _add_sheet(name: str, header: Sequence[str], rows: Iterable[Sequence[Any]]) -> None:
            ws = wb.create_sheet(title=name[:31])
            ws.append(list(header))
            for row in rows:
                ws.append(_rounded_row(row))

        _add_sheet("all_stims", header, rows)

        final_path = output_path if output_path.suffix.lower() == ".xlsx" else output_path.with_suffix(".xlsx")
        _ensure_parent(final_path)
        wb.save(final_path)
        created.append(final_path)
        return created

    raise ValueError(f"Unsupported file format: {file_format}")


def export_run_continuous_csv(
    *,
    run_data: Dict[str, Any],
    output_path: Path,
    selected_channels: List[int] | None = None,
    epoch_window_ms: Optional[float] = None,
    file_format: str = "csv",
    skip_pauses: bool = False,
) -> Path:
    """Непрерывная ЭЭГ за прогон → один файл.

    Колонки: ``sample_idx`` (0..N от начала записи), ``t_rel_s`` (секунды от первого отсчёта),
    ``ts`` (сырой штамп потока), ``ch_1``..``ch_N``, ``marker`` (0 — ни одна плитка не горит;
    иначе — номер плитки, которая горит **на протяжении всего отрезка** от её ``|on`` до
    ``|off`` из LSL-потока плиток), ``in_epoch`` (1 — отсчёт попадает в окно эпохи после
    какой-либо вспышки, 0 — пауза; окно берётся из ``epoch_window_ms`` либо
    ``epoch_time_ms``/шаблона, по умолчанию 800 мс).

    ``file_format``: ``csv`` (русский формат Excel: ``;`` разделитель и ``,`` дробная часть)
    или ``xlsx`` (нативные числа, Excel сам подставит запятую по локали).

    ``skip_pauses``: если True — в файл пишутся только отсчёты с ``in_epoch=1``
    (непрерывные «вспышечные» окна длиной ``epoch_window_ms`` после каждой вспышки).
    Тогда **первая строка файла = первая вспышка**, а паузы между вспышками
    не попадают в экспорт. Полезно, если в Excel не хочется листать «мёртвые» тысячи
    строк в начале записи.
    """
    fmt = (file_format or "csv").lower().strip()
    if fmt not in {"csv", "xlsx"}:
        raise ValueError(f"Unsupported file format for continuous export: {file_format}")

    data = _filtered_run_data(run_data, selected_channels)
    eeg_ts = data.get("eeg_ts") or []
    eeg_samples = data.get("eeg_samples") or []
    if not eeg_ts or not eeg_samples:
        raise RuntimeError("В прогоне нет непрерывных eeg_ts/eeg_samples — нечего экспортировать.")
    if len(eeg_ts) != len(eeg_samples):
        raise RuntimeError(
            f"Длины eeg_ts ({len(eeg_ts)}) и eeg_samples ({len(eeg_samples)}) не совпадают."
        )

    ts_arr = np.asarray(eeg_ts, dtype=np.float64)
    n = int(ts_arr.shape[0])

    summary = data.get("summary") or {}
    ap = summary.get("analysis_params") or {}

    def _f(x: Any) -> Optional[float]:
        try:
            if x is None:
                return None
            v = float(x)
            return v
        except (TypeError, ValueError):
            return None

    nominal_srate = _f(ap.get("sampling_rate_hz"))
    eff_srate_summary = _f(ap.get("sampling_rate_hz_effective"))
    est_srate_summary = _f(ap.get("sampling_rate_hz_estimated"))
    last_lc = _f(summary.get("lsl_clock_at_buffer_end"))
    first_lc = _f(summary.get("lsl_clock_at_buffer_start"))
    mk_offset = _f(summary.get("marker_eeg_offset"))
    ref_n = summary.get("lsl_clock_buffer_end_n_samples")
    try:
        ref_n_int = int(ref_n) if ref_n is not None else 0
    except (TypeError, ValueError):
        ref_n_int = 0
    if ref_n_int <= 0:
        ref_n_int = n

    # Оценка частоты: предпочтение — eff_srate из summary; затем nominal; затем первая/последняя lc.
    srate: Optional[float] = None
    for cand in (eff_srate_summary, nominal_srate, est_srate_summary):
        if cand is not None and cand > 0:
            srate = cand
            break
    if srate is None and last_lc is not None and first_lc is not None and last_lc > first_lc and n > 1:
        srate = float(n) / float(last_lc - first_lc)

    # t_rel_s: предпочитаем индекс/частоту (штамп NeuroSpectrum часто идёт шагом ~1 с),
    # иначе — по сырому ts.
    if srate and srate > 0:
        t_rel_s = np.arange(n, dtype=np.float64) / float(srate)
    else:
        t_rel_s = ts_arr - float(ts_arr[0]) if n > 0 else np.zeros(0, dtype=np.float64)

    # Привязка маркеров к индексам ЭЭГ. Приоритет:
    # 1) предвычисленный ``sample_idx`` на самом маркере (заполняется в qt_window на финализации);
    # 2) пересчёт на лету: marker_ts + offset → lsl_local_clock(), далее seconds_back * srate от ref_n;
    # 3) argmin |eeg_ts - marker_ts| — последний запасной вариант (ненадёжно при грубых штампах).
    can_lc_map = (
        last_lc is not None and srate is not None and srate > 0 and ref_n_int > 0
    )

    def _resolve_sample_idx(item: Dict[str, Any]) -> Optional[int]:
        """Сопоставляет маркеру (любому — on/off/trial_*) индекс отсчёта ЭЭГ."""
        pre = item.get("sample_idx")
        if isinstance(pre, (int, np.integer)) and not isinstance(pre, bool):
            cand = int(pre)
            if 0 <= cand < n:
                return cand
        if can_lc_map:
            try:
                ts_m = float(item.get("ts"))
            except (TypeError, ValueError):
                ts_m = None
            if ts_m is not None:
                t_mark_lc = ts_m + float(mk_offset) if mk_offset is not None else ts_m
                seconds_back = float(last_lc) - t_mark_lc
                cand = int(ref_n_int) - 1 - int(round(seconds_back * float(srate)))
                if 0 <= cand < n:
                    return cand
        try:
            ts_m = float(item.get("ts"))
        except (TypeError, ValueError):
            return None
        cand = int(np.argmin(np.abs(ts_arr - ts_m)))
        if 0 <= cand < n:
            return cand
        return None

    # Парсим события on/off в виде (sample_idx, tile, phase).
    events: List[Tuple[int, int, str]] = []
    for item in data.get("markers") or []:
        parsed = _parse_tile_event(item.get("value"))
        if parsed is None:
            continue
        tile, phase = parsed
        sidx = _resolve_sample_idx(item)
        if sidx is None:
            continue
        events.append((sidx, tile, phase))
    events.sort(key=lambda x: (x[0], 0 if x[2] == "on" else 1))

    # Заполняем marker: пока плитка «горит» (между её |on и |off) — её номер
    # на каждом отсчёте, иначе 0. Если |off не пришёл (короткий прогон,
    # остановка анализа в момент горения) — считаем, что плитка горит до конца записи.
    marker_vals = np.zeros(n, dtype=np.int64)
    flash_sample_idx: List[int] = []
    # Для каждой плитки храним последний sample_idx её |on, для которого не пришёл |off.
    open_on: Dict[int, int] = {}
    for sidx, tile, phase in events:
        if phase == "on":
            # Если предыдущая |on не закрылась — закрываем её «на себе» (следующая вспышка = новый «стартовый» момент).
            prev_on = open_on.get(tile)
            if prev_on is not None:
                end = min(n, sidx)
                marker_vals[prev_on:end] = 100 + tile
            open_on[tile] = sidx
            flash_sample_idx.append(sidx)
        elif phase == "off":
            start_on = open_on.pop(tile, None)
            if start_on is None:
                # |off без пары |on (до начала записи или потерянный) — ставим только точку.
                marker_vals[sidx] = 100 + tile
                continue
            end = min(n, sidx + 1)  # включаем сам отсчёт |off
            marker_vals[start_on:end] = 100 + tile
    # Хвост: всё, что осталось открытым — тянется до конца записи.
    for tile, start_on in open_on.items():
        marker_vals[start_on:n] = 100 + tile

    if epoch_window_ms is None:
        tpl = data.get("epoch_time_ms") or []
        if tpl:
            epoch_window_ms = float(tpl[-1])
        else:
            epoch_window_ms = 800.0

    win_samples: int
    if srate and srate > 0:
        win_samples = int(round(float(epoch_window_ms) / 1000.0 * srate))
    else:
        win_samples = int(round((n - 1) / max(1.0, float(t_rel_s[-1]) if n > 1 else 1.0)
                                * float(epoch_window_ms) / 1000.0))
    win_samples = max(1, win_samples)

    in_epoch = np.zeros(n, dtype=np.int64)
    for j in flash_sample_idx:
        in_epoch[j : min(n, j + win_samples)] = 1

    # Build per-sample target_tile_id: value of the last trial_start|target=N marker seen.
    # -1 means no trial has started yet; resets to -1 on trial_end.
    target_events: List[Tuple[int, int]] = []  # (sample_idx, target_id), -1 = trial_end
    for item in data.get("markers") or []:
        val = item.get("value")
        target = parse_trial_target_tile_id(val)
        if target is not None:
            sidx = _resolve_sample_idx(item)
            if sidx is not None:
                target_events.append((sidx, target))
        elif isinstance(val, str) and "trial_end" in val:
            sidx = _resolve_sample_idx(item)
            if sidx is not None:
                target_events.append((sidx, -1))
    target_events.sort(key=lambda x: x[0])

    target_tile_id_vals = np.full(n, -1, dtype=np.int64)
    current_target = -1
    ev_iter = iter(target_events)
    next_ev = next(ev_iter, None)
    for i in range(n):
        while next_ev is not None and next_ev[0] <= i:
            current_target = next_ev[1]
            next_ev = next(ev_iter, None)
        target_tile_id_vals[i] = current_target

    max_ch = max((len(x) for x in eeg_samples if isinstance(x, list)), default=0)
    if max_ch == 0:
        max_ch = 1
    stim = dict(summary.get("stimulus_params") or {})
    last_cue = summary.get("last_lsl_cue")
    ui_winner = summary.get("ui_winner_tile_id")
    run_seq = run_data.get("run_seq")
    saved_at_ms = run_data.get("saved_at_ms")
    const_meta = [
        run_seq,
        saved_at_ms,
        ap.get("baseline_ms"),
        ap.get("window_x_ms"),
        ap.get("window_y_ms"),
        ap.get("artifact_threshold_uv"),
        ap.get("use_car"),
        ap.get("epochs_after_trial_only"),
        last_cue,
        ui_winner,
        stim.get("target"),
        stim.get("sequences"),
        stim.get("isi_s"),
        stim.get("flash_s"),
        stim.get("cue_s"),
        stim.get("ready_s"),
        stim.get("inter_block_s"),
    ]

    header = ["sample_idx", "t_rel_s", "ts"] + [f"ch_{i+1}" for i in range(max_ch)] + [
        "run_seq",
        "saved_at_ms",
        "baseline_ms",
        "window_x_ms",
        "window_y_ms",
        "artifact_threshold_uv",
        "use_car",
        "epochs_after_trial_only",
        "last_lsl_cue",
        "ui_winner_tile_id",
        "stim_target",
        "stim_sequences",
        "stim_isi_s",
        "stim_flash_s",
        "stim_cue_s",
        "stim_ready_s",
        "stim_inter_block_s",
        "marker",
        "in_epoch",
        "target_tile_id",
    ]
    rows: List[List[Any]] = []
    for i, (t, smp, mk, ie, tgt) in enumerate(
        zip(eeg_ts, eeg_samples, marker_vals.tolist(), in_epoch.tolist(), target_tile_id_vals.tolist())
    ):
        if skip_pauses and int(ie) == 0:
            continue
        vals = smp if isinstance(smp, list) else [smp]
        row: List[Any] = [int(i), float(t_rel_s[i]), float(t)]
        for k in range(max_ch):
            row.append(vals[k] if k < len(vals) else None)
        row.extend(const_meta)
        row.append(int(mk))
        row.append(int(ie))
        row.append(int(tgt))
        rows.append(row)

    stem = output_path.with_suffix("")
    base_name = stem.name
    if base_name.endswith("_continuous"):
        base_name = base_name[: -len("_continuous")]
    if skip_pauses:
        base_name = f"{base_name}_flashes_only"
    else:
        base_name = f"{base_name}_continuous"
    target_dir = stem.parent
    if fmt == "xlsx":
        p = target_dir / f"{base_name}.xlsx"
        _rows_to_xlsx(p, "flashes" if skip_pauses else "continuous", header, rows)
    else:
        p = target_dir / f"{base_name}.csv"
        _rows_to_csv_ru(p, header, rows)
    return p
